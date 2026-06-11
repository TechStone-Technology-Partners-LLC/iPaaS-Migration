"""Generate gap Excel and Map field mapping Excel for webMethods GLD Compliance migration."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HDR_FONT   = Font(bold=True, color='FFFFFF')
HDR_FILL_B = PatternFill('solid', fgColor='1F4E79')   # blue header
HDR_FILL_G = PatternFill('solid', fgColor='375623')   # green header
WRAP       = Alignment(wrap_text=True, vertical='top')
ROW_A      = PatternFill('solid', fgColor='E2EFDA')
ROW_B      = PatternFill('solid', fgColor='FFFFFF')


def _hdr(ws, headers, fill):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = HDR_FONT
        c.fill = fill
        c.alignment = WRAP


# ─── 1. missing_components.xlsx ───────────────────────────────────────────────
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = 'Missing Components'

_hdr(ws1, ['WebMethods Component', 'Component Type', 'Description',
           'Boomi Equivalent', 'Mapping Notes'], HDR_FILL_B)

GAPS = [
    (
        'JDBC Adapter Connection (ConnectionData)',
        'Connection Node',
        'webMethods JDBC Adapter connection (GLDComplianceAdapterEnv:ExpressOS). '
        'Holds Oracle JDBC connection properties: host CSC06DSHORA1S, port 1522, '
        'SID ILMSUM, user GLD_SCHEMA, pool min=1/max=10.',
        'DatabaseV2 Connection (connector-settings)',
        'Mapped to MIG_WM_GLD_DB_Connection (ID 370bf544). '
        'JDBC URL: jdbc:oracle:thin:@CSC06DSHORA1S:1522/ILMSUM. '
        'Configure password via Boomi Environment Extensions.',
    ),
    (
        'Adapter Service - JDBC SELECT',
        'Adapter Service',
        'webMethods JDBC adapter service for SELECT queries '
        '(pub.db.jdbc:select or custom JDBC adapter operation).',
        'DatabaseV2 Operation (connector-action)',
        'Mapped to MIG_WM_GLD_QueryCompliance_Operation (ID 62cc118c). '
        'SQL: SELECT from GLD_SCHEMA.COMPLIANCE_RECORDS. '
        'Adjust column list and WHERE clause after import.',
    ),
    (
        'FLOW Service (pub.flow:service container)',
        'Flow Service Container',
        'Top-level webMethods IS flow service. '
        'Container for all steps; more granular than a Workflow.',
        'Process',
        'Each FLOW service maps to its own Boomi process. '
        'See MIG_WM_GLDCompliance_Process (ID 8c2d51b4). '
        'The Excel Workflow -> Process mapping covers the outer container only.',
    ),
    (
        'IS Document Type (node_type=record)',
        'Data Type / Schema',
        'webMethods IS typed document structure. '
        'Equivalent to a JSON/XML schema definition.',
        'JSON or XML Profile (profile.json / profile.xml)',
        'No IS document types found in GLDComplianceAdapterEnv. '
        'Create profiles in Boomi GUI and reference them in Map shapes.',
    ),
    (
        'ISMemDataImp (runtime pipeline object)',
        'Runtime Data Structure',
        'webMethods in-memory pipeline document '
        '(com.wm.data.ISMemDataImpl). Carries all data between steps.',
        'Document flowing through process + DDPs',
        'In Boomi, pipeline data flows as documents between shapes. '
        'Individual fields are extracted using Set Properties into DDPs.',
    ),
    (
        'ELSEIF (chained IF conditional)',
        'Flow Control',
        'webMethods ELSEIF block -- chained condition after an IF block.',
        'Chained Decision shapes on the false path',
        'The Excel maps ELSE and ELSEIF to Business Rules but does not '
        'detail the chaining pattern. Model as sequential Decision shapes '
        'on the false path of the preceding Decision.',
    ),
    (
        'CONTINUE (skip loop iteration)',
        'Flow Control',
        'webMethods CONTINUE statement inside a LOOP -- '
        'skip the current iteration and proceed to the next.',
        'dataContext.discard() in Groovy Data Process',
        'The Excel maps CONTINUE -> Loop but omits the skip-iteration detail. '
        'Use dataContext.discard() in the Groovy script inside the '
        'Data Process shape to skip a record. See shape7 in process.',
    ),
    (
        'pub.flow:sequence (built-in sequential container)',
        'Built-in Flow Service',
        'webMethods built-in service that wraps a sequence of steps. '
        'Functionally equivalent to an ordered list of invocations.',
        'Inline sequential steps OR Branch shape',
        'Overlaps with the SEQUENCE -> Branch mapping. '
        'For simple sequential ordering, inline steps in the main flow suffice; '
        'use Branch only when two distinct processing tracks are required.',
    ),
    (
        'Adapter Notification / pub.jms:send',
        'JMS / Messaging Service',
        'webMethods JMS send service -- publishes a message to a JMS destination.',
        'Boomi Event Streams (publish operation)',
        'Not present in GLDComplianceAdapterEnv but common in IS packages. '
        'Map to Event Streams Produce operation if migrating JMS-based flows.',
    ),
]

for i, row in enumerate(GAPS):
    ws1.append(row)
    fill = ROW_A if i % 2 == 0 else ROW_B
    for col in range(1, 6):
        c = ws1.cell(row=ws1.max_row, column=col)
        c.alignment = WRAP
        c.fill = fill

ws1.column_dimensions['A'].width = 38
ws1.column_dimensions['B'].width = 24
ws1.column_dimensions['C'].width = 52
ws1.column_dimensions['D'].width = 36
ws1.column_dimensions['E'].width = 62

wb1.save('WebMethods/missing_components.xlsx')
print(f'Written: WebMethods/missing_components.xlsx  ({len(GAPS)} rows)')


# ─── 2. map_field_mappings.xlsx ───────────────────────────────────────────────
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Map Field Mappings'

_hdr(ws2, ['Source (Pipeline In)', 'Transformation Logic',
           'Target (Pipeline Out)', 'Data Type', 'Notes'], HDR_FILL_G)

MAP_ROWS = [
    (
        'COMPLIANCE_RECORDS/RECORD_ID',
        'Direct copy',
        'complianceOutput/recordId',
        'STRING',
        'Primary key -- no transformation needed.',
    ),
    (
        'COMPLIANCE_RECORDS/ENTITY_ID',
        'Direct copy',
        'complianceOutput/entityId',
        'STRING',
        'Entity identifier.',
    ),
    (
        'COMPLIANCE_RECORDS/COMPLIANCE_STATUS',
        'Lookup translate: PASS -> Approved; FAIL -> Rejected; PENDING -> Under_Review',
        'complianceOutput/status',
        'STRING',
        'Status code translation. Implement as a Cross-Reference Table in Boomi.',
    ),
    (
        'COMPLIANCE_RECORDS/REVIEWED_BY',
        'Direct copy',
        'complianceOutput/reviewedBy',
        'STRING',
        'Reviewer name.',
    ),
    (
        'COMPLIANCE_RECORDS/REVIEW_DATE',
        'Date format: yyyy-MM-dd HH:mm:ss -> yyyy-MM-ddTHH:mm:ssZ (ISO-8601)',
        'complianceOutput/reviewDate',
        'DATETIME',
        'Use Boomi Map date/time function: CurrentDateTimeToFormat with target mask.',
    ),
    (
        '',
        '--- Template row: add actual MAP step pipeline fields below ---',
        '',
        '',
        'No MAP steps were found in GLDComplianceAdapterEnv (only a JDBC connection node). '
        'Populate when flow services with MAP steps are available.',
    ),
]

for i, row in enumerate(MAP_ROWS):
    ws2.append(row)
    fill = ROW_A if i % 2 == 0 else ROW_B
    for col in range(1, 6):
        c = ws2.cell(row=ws2.max_row, column=col)
        c.alignment = WRAP
        c.fill = fill

ws2.column_dimensions['A'].width = 42
ws2.column_dimensions['B'].width = 58
ws2.column_dimensions['C'].width = 42
ws2.column_dimensions['D'].width = 13
ws2.column_dimensions['E'].width = 58

# Instructions sheet
ws_info = wb2.create_sheet('Instructions')
ws_info['A1'] = 'How to use this file'
ws_info['A1'].font = Font(bold=True, size=12)
NOTES = [
    '',
    'This file documents MAP shape field mappings for the webMethods -> Boomi GLD Compliance migration.',
    '',
    'Column descriptions:',
    '  Source (Pipeline In)     — Field path from the webMethods pipeline input document.',
    '  Transformation Logic     — Script, formula, lookup, or "Direct copy" used to derive output.',
    '  Target (Pipeline Out)    — Field path in the Boomi output document / JSON profile.',
    '  Data Type                — webMethods IS data type (STRING, INTEGER, DATETIME, OBJECT, etc.)',
    '  Notes                    — Edge cases, TODO items, or Boomi Map function references.',
    '',
    'Source package: GLDComplianceAdapterEnv (webMethods IS 6.5, keybank.com, 2008-03-27)',
    'No MAP steps were found in the analyzed package (only a JDBC adapter connection node).',
    'The rows in the Map Field Mappings sheet are representative placeholders.',
    'Populate them when flow services containing MAP steps are provided.',
    '',
    'Boomi Map documentation: https://help.boomi.com/docs/atomsphere/integration/process-building/r-atm-map_shape_0abe324f-c3b6-437c-aeac-b6c1c7d5d0a4/',
]
for note in NOTES:
    ws_info.append([note])
ws_info.column_dimensions['A'].width = 90

wb2.save('WebMethods/map_field_mappings.xlsx')
print(f'Written: WebMethods/map_field_mappings.xlsx  ({len(MAP_ROWS)} data rows)')
