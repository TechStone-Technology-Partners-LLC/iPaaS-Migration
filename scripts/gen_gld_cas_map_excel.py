"""Update map_field_mappings.xlsx with actual GLDComplianceAdapterServices pipeline fields."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HDR_FONT   = Font(bold=True, color='FFFFFF')
HDR_FILL_G = PatternFill('solid', fgColor='375623')
WRAP       = Alignment(wrap_text=True, vertical='top')
ROW_A      = PatternFill('solid', fgColor='E2EFDA')
ROW_B      = PatternFill('solid', fgColor='FFFFFF')

def _hdr(ws, headers, fill):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = HDR_FONT
        c.fill = fill
        c.alignment = WRAP

wb = openpyxl.Workbook()

# ── Sheet 1: logCheckRequest Input Mapping ────────────────────────────────────
ws1 = wb.active
ws1.title = 'logCheckRequest'
_hdr(ws1, ['Source (Pipeline In)', 'Transformation Logic', 'Target (Pipeline Out)', 'Data Type', 'Notes'], HDR_FILL_G)

ROWS1 = [
    ('request/CustomerNbr',           'Direct copy', 'CUSTOMERNBR',           'VARCHAR2(18)',  ''),
    ('request/CustomerType',           'Direct copy', 'CUSTOMERTYPE',           'VARCHAR2(3)',   ''),
    ('request/PartyType',              'Direct copy', 'PARTYTYPE',              'VARCHAR2(20)',  ''),
    ('request/Businessname',           'Direct copy', 'BUSINESSNAME',           'VARCHAR2(40)',  ''),
    ('request/ApplicationNbr',         'Direct copy', 'APPLICATIONNBR',         'VARCHAR2(18)',  ''),
    ('request/Channel',                'Direct copy', 'CHANNEL',                'VARCHAR2(10)',  ''),
    ('request/LOB',                    'Direct copy', 'LOB',                    'VARCHAR2(10)',  ''),
    ('request/ProductCode',            'Direct copy', 'PRODUCTCODE',            'VARCHAR2(10)',  ''),
    ('request/SubProductCode',         'Direct copy', 'SUBPRODUCTCODE',         'VARCHAR2(10)',  ''),
    ('request/PostBack',               'Direct copy', 'POSTBACK',               'VARCHAR2(200)', ''),
    ('request/ComplianceReplyEmail',   'Direct copy', 'COMPLIANCEREPLYEMAIL',   'VARCHAR2(75)',  ''),
    ('request/FirstName',              'Direct copy', 'FIRSTNAME',              'VARCHAR2(20)',  ''),
    ('request/MiddleName',             'Direct copy', 'MIDDLENAME',             'VARCHAR2(20)',  ''),
    ('request/LastName',               'Direct copy', 'LASTNAME',               'VARCHAR2(20)',  ''),
    ('request/AddressLine1',           'Direct copy', 'ADDRESSLINE1',           'VARCHAR2(40)',  ''),
    ('request/AddressLine2',           'Direct copy', 'ADDRESSLINE2',           'VARCHAR2(40)',  ''),
    ('request/AddressLine3',           'Direct copy', 'ADDRESSLINE3',           'VARCHAR2(40)',  ''),
    ('request/AddressLine4',           'Direct copy', 'ADDRESSLINE4',           'VARCHAR2(40)',  ''),
    ('request/City',                   'Direct copy', 'CITY',                   'VARCHAR2(20)',  ''),
    ('request/State',                  'Direct copy', 'STATE',                  'VARCHAR2(2)',   '2-char state code'),
    ('request/Zip',                    'Direct copy', 'ZIP',                    'VARCHAR2(10)',  ''),
    ('request/CountryCode',            'Direct copy', 'COUNTRYCODE',            'VARCHAR2(3)',   '3-char ISO code'),
    ('request/SSNTIN',                 'Direct copy (PII — do not log)', 'SSNTIN', 'VARCHAR2(9)', 'SSN or TIN. Set enableUserLog=false on process.'),
    ('request/DOB',                    'Date format: input yyyy-MM-dd → Oracle DATE', 'DOB', 'DATE', 'Use Boomi Map CurrentDateTimeToFormat function'),
    ('request/RequestorSystemRequestID', 'Convert String to Long', 'REQUESTORSYSTEMREQUESTID', 'BIGINT', 'Cast using Boomi Map Number function'),
    ('(OUT) Oracle auto-generated',    'Returned by SP', 'accCheckRequestID', 'Long/BIGINT', 'Store as DPP_ACC_CHECK_REQUEST_ID for later use'),
]
for i, row in enumerate(ROWS1):
    ws1.append(row)
    fill = ROW_A if i % 2 == 0 else ROW_B
    for col in range(1, 6):
        c = ws1.cell(row=ws1.max_row, column=col)
        c.alignment = WRAP
        c.fill = fill
ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 50
ws1.column_dimensions['C'].width = 30
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 50

# ── Sheet 2: selectCustomerAndRequest Output Mapping ─────────────────────────
ws2 = wb.create_sheet('selectCustomerAndRequest')
_hdr(ws2, ['Source (Pipeline In)', 'Transformation Logic', 'Target (Pipeline Out)', 'Data Type', 'Notes'], HDR_FILL_G)

ROWS2 = [
    ('DPP_CIU_REF_NBR',                'Direct copy → SQL WHERE param', 'WHERE t2.CIUREFNBR = ?', 'VARCHAR2(20)', 'Set before calling this operation'),
    ('results[]/ACCCUSTOMERID',         'Direct copy', 'DPP_ACC_CUSTOMER_ID',   'BigDecimal', ''),
    ('results[]/CUSTOMERNBR',           'Direct copy', 'DPP_CUSTOMER_NBR',      'String',     ''),
    ('results[]/CUSTOMERTYPE',          'Direct copy', 'DPP_CUSTOMER_TYPE',     'String',     ''),
    ('results[]/BUSINESSNAME',          'Direct copy', 'DPP_BUSINESS_NAME',     'String',     ''),
    ('results[]/FIRSTNAME',             'Direct copy', 'DPP_FIRST_NAME',        'String',     ''),
    ('results[]/MIDDLENAME',            'Direct copy', 'DPP_MIDDLE_NAME',       'String',     ''),
    ('results[]/LASTNAME',              'Direct copy', 'DPP_LAST_NAME',         'String',     ''),
    ('results[]/ADDRESSLINE1',          'Direct copy', 'DPP_ADDRESS_LINE1',     'String',     ''),
    ('results[]/ADDRESSLINE2',          'Direct copy', 'DPP_ADDRESS_LINE2',     'String',     ''),
    ('results[]/CITY',                  'Direct copy', 'DPP_CITY',              'String',     ''),
    ('results[]/STATE',                 'Direct copy', 'DPP_STATE',             'String',     ''),
    ('results[]/ZIP',                   'Direct copy', 'DPP_ZIP',               'String',     ''),
    ('results[]/SSNTIN',                'Direct copy (PII)', 'DPP_SSNTIN',      'String',     'Handle as sensitive data'),
    ('results[]/DOB',                   'Timestamp → String yyyy-MM-dd', 'DPP_DOB', 'Timestamp', ''),
    ('results[]/PARTYTYPE',             'Direct copy', 'DPP_PARTY_TYPE',        'String',     ''),
    ('results[]/ACCCHECKREQUESTID',     'Direct copy', 'DPP_ACC_CHECK_REQUEST_ID', 'BigDecimal', ''),
    ('results[]/APPLICATIONNBR',        'Direct copy', 'DPP_APPLICATION_NBR',   'String',     ''),
    ('results[]/CHANNEL',               'Direct copy', 'DPP_CHANNEL',           'String',     ''),
    ('results[]/LOB',                   'Direct copy', 'DPP_LOB',               'String',     ''),
    ('results[]/PRODUCTCODE',           'Direct copy', 'DPP_PRODUCT_CODE',      'String',     ''),
    ('results[]/SUBPRODUCTCODE',        'Direct copy', 'DPP_SUB_PRODUCT_CODE',  'String',     ''),
    ('results[]/COMPLIANCEREPLYEMAIL',  'Direct copy', 'DPP_COMPLIANCE_REPLY_EMAIL', 'String', ''),
    ('results[]/CIUREFNBR',             'Direct copy', 'DPP_CIU_REF_NBR',       'String',     'Confirm matches input'),
    ('results[]/REQUESTTIMESTAMP',      'Timestamp → ISO-8601 String', 'DPP_REQUEST_TIMESTAMP', 'Timestamp', 'Use CurrentDateTimeToFormat'),
]
for i, row in enumerate(ROWS2):
    ws2.append(row)
    fill = ROW_A if i % 2 == 0 else ROW_B
    for col in range(1, 6):
        c = ws2.cell(row=ws2.max_row, column=col)
        c.alignment = WRAP
        c.fill = fill
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 44
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 44

# ── Sheet 3: logCheckReplyError Mapping ───────────────────────────────────────
ws3 = wb.create_sheet('logCheckReplyError')
_hdr(ws3, ['Source (Pipeline In)', 'Transformation Logic', 'Target (Pipeline Out)', 'Data Type', 'Notes'], HDR_FILL_G)

ROWS3 = [
    ('$error/errorMessage', 'Direct copy',            'ERRORDESC',  'String', 'webMethods $error pipeline variable'),
    ('$error/errorCode',    'Direct copy',            'ERRORCODE',  'String', ''),
    ('$error/errorType',    'Direct copy',            'ERRORTYPE',  'String', ''),
    ('DPP_CIU_REF_NBR',    'Direct copy',            'CIUREFNBR',  'String', 'Must be set before entering error path'),
]
for i, row in enumerate(ROWS3):
    ws3.append(row)
    fill = ROW_A if i % 2 == 0 else ROW_B
    for col in range(1, 6):
        c = ws3.cell(row=ws3.max_row, column=col)
        c.alignment = WRAP
        c.fill = fill
ws3.column_dimensions['A'].width = 36
ws3.column_dimensions['B'].width = 36
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 50

# ── Sheet 4: Instructions ─────────────────────────────────────────────────────
ws_info = wb.create_sheet('Instructions')
ws_info['A1'] = 'How to use this file'
ws_info['A1'].font = Font(bold=True, size=12)
NOTES = [
    '',
    'Package: GLDComplianceAdapterServices (webMethods IS 6.5, Oracle JDBC adapter)',
    'Generated from node.ndf analysis of all 7 adapter services.',
    '',
    'Sheets:',
    '  logCheckRequest           — Maps input request document fields to ACCLOGCHECKREQUEST SP parameters',
    '  selectCustomerAndRequest  — Maps CIURefNbr input and output row fields from the JOIN SELECT query',
    '  logCheckReplyError        — Maps error pipeline to ACCLOGCHECKREPLYERROR SP parameters',
    '',
    'Sheets not included (direct-copy or void SPs):',
    '  logCheckReply      — IN: CIURefNbr (String), CheckType (String), Result (Boolean→VARCHAR2)',
    '  logCheckRequestXML — IN: ApplicationID (Long), Request (CLOB), RequestIdentifier1/2/3 (String)',
    '  updateCIURefNbr    — IN: accCheckRequestID (Long), CIURefNbr (String)',
    '  purgeData          — No inputs or outputs',
    '',
    'Column descriptions:',
    '  Source (Pipeline In)    — webMethods pipeline variable or DB result field',
    '  Transformation Logic    — "Direct copy", date format, type cast, or formula',
    '  Target (Pipeline Out)   — Boomi DDP name or DB column/SP parameter',
    '  Data Type               — Oracle or Java data type',
    '  Notes                   — PII flags, format requirements, Boomi Map function hints',
]
for note in NOTES:
    ws_info.append([note])
ws_info.column_dimensions['A'].width = 90

wb.save('WebMethods/Analysis/map_field_mappings.xlsx')
print('Written: WebMethods/Analysis/map_field_mappings.xlsx — 3 data sheets + Instructions')
